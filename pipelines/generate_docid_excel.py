"""
Export DocID ↔ Book Name Mapping to Excel
Auto-appends newly ingested books.
"""

import os
from openpyxl import Workbook, load_workbook

from core.registry.document_registry import DocumentRegistry
from core.utils.logging_utils import get_component_logger


# =====================================================
# LOGGER SETUP
# =====================================================

logger = get_component_logger("DocIDExporter", component="ingestion")


OUTPUT_FILE = "docid_book_mapping.xlsx"


def main():

    try:
        registry = DocumentRegistry()
        rows = registry.fetch_all()

        if not rows:
            logger.warning("No records found in registry.")
            return

        # =====================================================
        # If file exists → append
        # If not → create new workbook
        # =====================================================

        if os.path.exists(OUTPUT_FILE):
            logger.info("Existing Excel found. Appending new records...")
            wb = load_workbook(OUTPUT_FILE)
            ws = wb.active

            existing_doc_ids = {
                row[0].value for row in ws.iter_rows(min_row=2)
            }

            new_count = 0

            for row in rows:
                if row[0] not in existing_doc_ids:
                    ws.append(row)
                    new_count += 1

            wb.save(OUTPUT_FILE)

            logger.info(f"Appended {new_count} new records.")

        else:
            logger.info("Creating new Excel file...")

            wb = Workbook()
            ws = wb.active
            ws.title = "DocID Mapping"

            ws.append(
                ["Doc ID", "Title", "Source Path", "Total Pages", "Created At"]
            )

            for row in rows:
                ws.append(row)

            wb.save(OUTPUT_FILE)

            logger.info(f"Created new file with {len(rows)} records.")

    except Exception:
        logger.exception("Failed exporting DocID mapping to Excel")
        raise


if __name__ == "__main__":
    main()
